"""Tests for conversation-printer: scope, format, filename, content, export."""

from pathlib import Path

from finops_buddy.agent.conversation_printer import (
    FORMAT_CSV,
    FORMAT_PDF,
    FORMAT_TXT,
    FORMAT_XLSX,
    SCOPE_FULL,
    SCOPE_LAST,
    SCOPE_QA,
    SCOPE_SUMMARY,
    build_content_for_scope,
    export_csv,
    export_pdf,
    export_pdf_fallback,
    export_txt,
    export_xlsx_from_table,
    export_xlsx_openpyxl,
    export_xlsx_via_mcp,
    generate_output_filename,
    get_output_path,
    parse_format_input,
    parse_scope_input,
    run_print_flow,
)


def test_parse_scope_input_full():
    """Scope 1 or 'full' returns SCOPE_FULL."""
    assert parse_scope_input("1") == SCOPE_FULL
    assert parse_scope_input("full") == SCOPE_FULL
    assert parse_scope_input("  FULL  ") == SCOPE_FULL


def test_parse_scope_input_qa():
    """Scope 2 or 'qa' returns SCOPE_QA."""
    assert parse_scope_input("2") == SCOPE_QA
    assert parse_scope_input("qa") == SCOPE_QA


def test_parse_scope_input_summary():
    """Scope 3 or 'summary' returns SCOPE_SUMMARY."""
    assert parse_scope_input("3") == SCOPE_SUMMARY
    assert parse_scope_input("summary") == SCOPE_SUMMARY


def test_parse_scope_input_last():
    """Scope 4 or 'last' returns SCOPE_LAST."""
    assert parse_scope_input("4") == SCOPE_LAST
    assert parse_scope_input("last") == SCOPE_LAST


def test_parse_scope_input_unknown_returns_none():
    """Unknown scope returns None."""
    assert parse_scope_input("x") is None
    assert parse_scope_input("") is None


def test_parse_format_input_txt_pdf_csv_xlsx():
    """Format parsing returns correct format id."""
    assert parse_format_input("txt") == FORMAT_TXT
    assert parse_format_input("1") == FORMAT_TXT
    assert parse_format_input("pdf") == FORMAT_PDF
    assert parse_format_input("2") == FORMAT_PDF
    assert parse_format_input("csv") == FORMAT_CSV
    assert parse_format_input("3") == FORMAT_CSV
    assert parse_format_input("xlsx") == FORMAT_XLSX
    assert parse_format_input("4") == FORMAT_XLSX


def test_parse_format_input_unknown_returns_none():
    """Unknown format returns None."""
    assert parse_format_input("doc") is None
    assert parse_format_input("") is None


def test_build_content_for_scope_full():
    """Full scope returns all turns as plain, markdown, and rows."""
    conv = ["User: Hello", "Agent: Hi there"]
    plain, md, rows = build_content_for_scope(conv, SCOPE_FULL)
    assert "Hello" in plain and "Hi there" in plain
    assert "## User" in md and "## Agent" in md
    assert len(rows) == 2
    assert rows[0] == (1, "User", "Hello")
    assert rows[1] == (1, "Agent", "Hi there")


def test_build_content_for_scope_last():
    """Last scope returns only the last agent message."""
    conv = ["User: Q1", "Agent: A1", "User: Q2", "Agent: A2"]
    plain, md, rows = build_content_for_scope(conv, SCOPE_LAST)
    assert "A2" in plain
    assert "A1" not in plain
    assert len(rows) == 1
    assert rows[0][2] == "A2"


def test_build_content_for_scope_summary():
    """Summary scope with summary_text uses that as content."""
    conv = ["User: Hi", "Agent: Hello"]
    summary = "User said hi and agent replied."
    plain, md, rows = build_content_for_scope(conv, SCOPE_SUMMARY, summary_text=summary)
    assert plain == summary
    assert md == summary
    assert len(rows) == 1
    assert rows[0][2] == summary


def test_generate_output_filename_uses_timestamp_and_slug():
    """Filename is timestamp + slug from first user message + ext."""
    conv = ["User: cost review for Q1", "Agent: Here are the costs..."]
    name = generate_output_filename(conv, "pdf")
    assert name.endswith(".pdf")
    assert "cost-review" in name or "cost" in name
    assert name[0:8].isdigit() and name[8] == "_"


def test_generate_output_filename_default_title_when_no_user_message():
    """When no user message, title is default."""
    conv = ["Agent: Some response"]
    name = generate_output_filename(conv, "txt")
    assert name.endswith(".txt")
    assert "conversation" in name


def test_get_output_path_uses_cwd_by_default():
    """Output path is under cwd when output_dir is None."""
    p = get_output_path("20250101_120000-chat.txt", output_dir=None)
    assert p.name == "20250101_120000-chat.txt"
    assert p.parent == Path.cwd()


def test_get_output_path_uses_given_dir():
    """Output path is under given directory."""
    p = get_output_path("out.csv", output_dir=Path("/tmp"))
    assert p.name == "out.csv"
    assert p.parent == Path("/tmp")


def test_export_txt_writes_content(tmp_path):
    """TXT export writes plain text to path."""
    export_txt("Hello\n\nWorld", tmp_path / "out.txt")
    assert (tmp_path / "out.txt").read_text(encoding="utf-8") == "Hello\n\nWorld"


def test_export_csv_writes_rows(tmp_path):
    """CSV export writes header and rows."""
    rows = [(1, "User", "Hi"), (2, "Agent", "Hello")]
    export_csv(rows, tmp_path / "out.csv")
    content = (tmp_path / "out.csv").read_text(encoding="utf-8")
    assert "turn,role,content" in content
    assert "Hi" in content and "Hello" in content


def test_run_print_flow_empty_conversation_prints_message(capsys):
    """When conversation is empty, run_print_flow prints no-conversation message."""
    mock_agent = object()
    out_lines = []

    def fake_print(*args):
        out_lines.append(" ".join(str(a) for a in args))

    run_print_flow(
        [],
        mock_agent,
        input_func=lambda: "",
        print_func=fake_print,
    )
    assert any("No conversation" in line for line in out_lines)


def test_pdf_export_when_disabled_returns_clear_message():
    """When PDF MCP is disabled, export_pdf_via_mcp returns (False, message)."""
    from unittest.mock import patch

    from finops_buddy.agent.conversation_printer import export_pdf_via_mcp

    with patch("finops_buddy.settings.get_pdf_mcp_enabled", return_value=False):
        ok, msg = export_pdf_via_mcp("# Hi", Path("/tmp/out.pdf"))
    assert ok is False
    assert "disabled" in msg.lower() or "PDF" in msg


def test_export_pdf_when_mcp_fails_uses_fallback(tmp_path):
    """When PDF MCP fails (e.g. stdout pollution), export_pdf uses weasyprint fallback."""
    from unittest.mock import patch

    path = tmp_path / "out.pdf"
    with patch("finops_buddy.settings.get_pdf_mcp_enabled", return_value=True):
        with patch(
            "finops_buddy.agent.conversation_printer.export_pdf_via_mcp",
            return_value=(False, "PDF export failed: unhandled errors in a TaskGroup"),
        ):
            ok, msg = export_pdf("# Summary\n\nSome **bold** text.", path)
    # Fallback (weasyprint) may succeed or fail depending on env (e.g. Cairo on Windows)
    if ok:
        assert path.exists()
        assert path.stat().st_size > 0
    else:
        assert "fallback" in msg.lower() or "weasyprint" in msg.lower() or "markdown" in msg.lower()


def test_export_pdf_fallback_creates_pdf_when_available(tmp_path):
    """export_pdf_fallback creates a PDF file when weasyprint is available."""
    path = tmp_path / "fallback.pdf"
    ok, msg = export_pdf_fallback("# Hello\n\nWorld.", path)
    if ok:
        assert path.exists()
        assert path.stat().st_size > 0
    else:
        # CI or env without weasyprint/Cairo
        assert "weasyprint" in msg.lower() or "markdown" in msg.lower() or "fallback" in msg.lower()


def test_excel_export_openpyxl_writes_file(tmp_path):
    """export_xlsx_openpyxl writes xlsx with header and rows."""
    rows = [(1, "User", "Hi"), (2, "Agent", "Hello")]
    path = tmp_path / "out.xlsx"
    export_xlsx_openpyxl(rows, path)
    assert path.exists()
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    assert ws is not None
    assert list(ws.iter_rows(max_row=3, values_only=True)) == [
        ("turn", "role", "content"),
        (1, "User", "Hi"),
        (2, "Agent", "Hello"),
    ]


def test_excel_export_when_mcp_disabled_uses_openpyxl(tmp_path):
    """When Excel MCP is disabled, export_xlsx_via_mcp uses openpyxl and succeeds."""
    from unittest.mock import patch

    with patch("finops_buddy.settings.get_excel_mcp_enabled", return_value=False):
        ok, msg = export_xlsx_via_mcp([(1, "User", "Hi")], tmp_path / "out.xlsx")
    assert ok is True
    assert (tmp_path / "out.xlsx").exists()


def test_excel_export_fallback_when_mcp_has_no_tools(tmp_path):
    """When MCP is enabled but lacks create_workbook/write_worksheet, openpyxl fallback succeeds."""
    from unittest.mock import patch

    with patch("finops_buddy.settings.get_excel_mcp_enabled", return_value=True):
        with patch(
            "finops_buddy.agent.conversation_printer._try_excel_mcp",
            return_value=(False, "Excel MCP does not expose create_workbook/write_worksheet."),
        ):
            ok, msg = export_xlsx_via_mcp([(1, "User", "Hi")], tmp_path / "out.xlsx")
    assert ok is True
    assert (tmp_path / "out.xlsx").exists()


def test_export_xlsx_from_table_writes_arbitrary_rows(tmp_path):
    """export_xlsx_from_table writes any list of rows (e.g. Q&A table) to xlsx."""
    data = [
        ["Question", "Answer"],
        ["What is FinOps?", "Cloud financial management."],
        ["Cost by service?", "Use cost-explorer tool."],
    ]
    path = tmp_path / "qa.xlsx"
    export_xlsx_from_table(data, path, sheet_name="Q&A")
    assert path.exists()
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    assert ws is not None
    assert ws.title == "Q&A"
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    assert rows == [tuple(r) for r in data]


def test_create_export_tools_returns_export_to_pdf_and_excel():
    """create_export_tools returns two tools callable by the agent."""
    from finops_buddy.agent.tools import create_export_tools

    tools = create_export_tools()
    assert len(tools) == 2
    names = [getattr(t, "__name__", str(t)) for t in tools]
    assert "export_to_pdf" in names
    assert "export_to_excel" in names


def test_export_to_excel_tool_writes_file(tmp_path):
    """Agent export_to_excel tool writes structured data to xlsx (e.g. Q&A table)."""
    import os

    from finops_buddy.agent.tools import create_export_tools

    tools = create_export_tools()
    export_to_excel = next(t for t in tools if getattr(t, "__name__", "") == "export_to_excel")
    old_cwd = os.getcwd()
    try:
        os.chdir(tmp_path)
        result = export_to_excel(
            [["Topic", "Summary"], ["Cost review", "User asked about Q1 costs."]],
            "agent_export.xlsx",
            "Summary",
        )
        assert "agent_export.xlsx" in result or (tmp_path / "agent_export.xlsx").exists()
        assert (tmp_path / "agent_export.xlsx").exists()
    finally:
        os.chdir(old_cwd)
